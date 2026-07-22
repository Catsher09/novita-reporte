import os
import json
import requests
import schedule
import time
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

FIREBASE_PROJECT = "novedades-campo-data"
FIREBASE_KEY = "AIzaSyDkyPmHf4-4tqldnCAxuOY4n_zNyp2kglI"
TELEGRAM_TOKEN = "8662511445:AAHe-SCgHwfY0FPFcKcOaCY0rJs25-tYBZE"
BASE = f"https://firestore.googleapis.com/v1/projects/{FIREBASE_PROJECT}/databases/(default)/documents"

RESOLUTORAS = {
    "daniela": {"nombre": "Daniela", "telegram_id": "7153477506"},
    "wendy":   {"nombre": "Wendy",   "telegram_id": None},  # se agrega después
}

def from_fs(doc):
    if not doc or "fields" not in doc:
        return None
    out = {"_id": doc["name"].split("/")[-1]}
    for k, v in doc["fields"].items():
        if "stringValue" in v: out[k] = v["stringValue"]
        elif "integerValue" in v: out[k] = int(v["integerValue"])
        elif "booleanValue" in v: out[k] = v["booleanValue"]
        else: out[k] = None
    return out

def fs_list(col):
    try:
        r = requests.get(f"{BASE}/{col}?key={FIREBASE_KEY}&pageSize=500", timeout=15)
        if not r.ok: return []
        data = r.json()
        return [from_fs(d) for d in data.get("documents", []) if from_fs(d)]
    except: return []

def fs_get(path):
    try:
        r = requests.get(f"{BASE}/{path}?key={FIREBASE_KEY}", timeout=15)
        return r.json() if r.ok else None
    except: return None

def get_mes_anterior():
    hoy = datetime.now()
    primer_dia_mes_actual = hoy.replace(day=1)
    ultimo_dia_mes_anterior = primer_dia_mes_actual - timedelta(days=1)
    return ultimo_dia_mes_anterior.month, ultimo_dia_mes_anterior.year

def parse_fecha(f):
    try:
        d, m, y = f.split("/")
        return datetime(int(y), int(m), int(d))
    except: return None

def generar_excel_mes(chats, mensajes, asignaciones, mes, anio):
    nombre_mes = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
                  "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"][mes-1]
    
    # Filtrar chats del mes
    chats_mes = []
    for c in chats:
        fc = parse_fecha(c.get("fecha",""))
        if fc and fc.month == mes and fc.year == anio:
            chats_mes.append(c)
    
    if not chats_mes:
        return None, nombre_mes
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"{nombre_mes} {anio}"
    
    # Encabezados
    headers = ["Fecha","Gestor","Colegio","Auditor","Resolutora","¿Tiene novedad?","Estado","Conversación completa"]
    header_fill = PatternFill(start_color="1d4ed8", end_color="1d4ed8", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Anchos de columna
    anchos = [12, 25, 30, 20, 15, 15, 10, 80]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[chr(64+i)].width = ancho
    
    # Datos
    for row, c in enumerate(chats_mes, 2):
        msgs = [m for m in mensajes if m.get("chatId") == c.get("_id")]
        lineas = [f"[{m.get('hora','')}] {m.get('autor','')}: {m.get('texto','')}" for m in msgs]
        
        resolutora_id = asignaciones.get(c.get("gestor",""), "")
        resolutora_nombre = RESOLUTORAS.get(resolutora_id, {}).get("nombre", "Sin asignar") if resolutora_id else "Sin asignar"
        novedad = c.get("novedad","")
        novedad_txt = "Sí" if novedad == "si" else "No" if novedad == "no" else "Sin marcar"
        estado = "Cerrado" if c.get("estado") == "cerrado" else "Abierto"
        
        fila = [c.get("fecha",""), c.get("gestor",""), c.get("colegio",""), c.get("auditorNombre",""), resolutora_nombre, novedad_txt, estado, " | ".join(lineas)]
        for col, val in enumerate(fila, 1):
            ws.cell(row=row, column=col, value=val)
        
        # Alternar colores de filas
        if row % 2 == 0:
            fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = fill
    
    filename = f"/tmp/Novita_Reporte_{nombre_mes}_{anio}.xlsx"
    wb.save(filename)
    return filename, nombre_mes

def enviar_reporte():
    print(f"[{datetime.now()}] Generando reporte mensual...")
    mes, anio = get_mes_anterior()
    
    # Obtener datos de Firebase
    chats = fs_list("chats")
    mensajes = fs_list("mensajes")
    config = fs_get("config/gestores")
    
    asignaciones = {}
    if config and "fields" in config:
        asig_str = config["fields"].get("asignaciones", {}).get("stringValue", "{}")
        try: asignaciones = json.loads(asig_str)
        except: pass
    
    archivo, nombre_mes = generar_excel_mes(chats, mensajes, asignaciones, mes, anio)
    
    if not archivo:
        print(f"Sin datos para {nombre_mes} {anio}")
        return
    
    # Enviar a cada resolutora
    for rid, info in RESOLUTORAS.items():
        if not info.get("telegram_id"):
            continue
        try:
            # Enviar mensaje previo
            requests.post(
                f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
                json={"chat_id": info["telegram_id"], "text": f"📊 Hola {info['nombre']}! Aquí está el reporte de <b>{nombre_mes} {anio}</b> con todos los chats del mes.", "parse_mode": "HTML"},
                timeout=10
            )
            # Enviar archivo Excel
            with open(archivo, "rb") as f:
                requests.post(
                    f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument",
                    data={"chat_id": info["telegram_id"]},
                    files={"document": (f"Novita_Reporte_{nombre_mes}_{anio}.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                    timeout=30
                )
            print(f"Reporte enviado a {info['nombre']}")
        except Exception as e:
            print(f"Error enviando a {info['nombre']}: {e}")

def run():
    print("Novita Reporte Server iniciado ✅")
    # Ejecutar el día 1 de cada mes a las 7am
    schedule.every().day.at("07:00").do(lambda: enviar_reporte() if datetime.now().day == 1 else None)
    
    # Para prueba: ejecutar también a los 5 minutos de iniciar si se pasa ?test=1
    if os.environ.get("TEST_MODE") == "1":
        print("Modo prueba: enviando reporte en 10 segundos...")
        time.sleep(10)
        enviar_reporte()
    
    while True:
        schedule.run_pending()
        time.sleep(60)

if __name__ == "__main__":
    run()
